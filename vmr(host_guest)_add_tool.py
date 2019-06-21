from API import Get_put_post
import openpyxl
from collections import OrderedDict
from profiles_passcode_check import decorator_check_profile_ids
import configparser

import logging


logging.basicConfig(filename="systemlogs.txt",
                    format='levelname:%(levelname)s filename: %(filename)s '
                           'outputNumber: [%(lineno)d]  func: %(funcName)s output msg:  %(message)s'
                           ' - %(asctime)s', datefmt='[%d/%b/%Y %H:%M:%S]', level=logging.INFO)

def read_serverinfo():
    cf = configparser.ConfigParser()
    cf.read("system_init_info.ini")
    server_ip = dict(cf.items("server_ip"))
    login = dict(cf.items("server_login"))
    return (server_ip, login)


def open_xslsheet():
    try:
        wb = openpyxl.load_workbook('./VMR_list.xlsx')
        ws=wb.get_sheet_by_name('add_host_guest')
    except Exception as error:
        logging.error(error)
    filerow=3
    while ws.cell(row=filerow,column=1).value:
        filerow +=1
    return (ws,filerow)


class Add_vmrs():
    def __init__(self,worksheet,api):
        self.ws = worksheet
        self.api=api

    def get_titles(self):
        c=1
        self.c_f={};self.vmr_features={};self.vmr_hostfeatures={};self.vmr_guestfeatures={};self.c_f[c]={}
        self.c_f=OrderedDict(self.c_f);self.vmr_features=OrderedDict(self.vmr_features)
        self.vmr_hostfeatures=OrderedDict(self.vmr_hostfeatures);self.vmr_guestfeatures=OrderedDict(self.vmr_guestfeatures)
        self.c_f=OrderedDict(self.c_f)
        while self.ws.cell(row=2,column=c).value:
            cellvalue= self.ws.cell(row=2,column=c).value
            self.c_f[c]=cellvalue
            if c==1:
                self.vmr_features['Vmr Number']=cellvalue
                self.c_f[c] = cellvalue
            elif (cellvalue.startswith('host')):
                self.vmr_hostfeatures[cellvalue[4:]] = None
                self.c_f[c] = cellvalue[4:]
                self.hostc = c
            else:
                self.vmr_guestfeatures[cellvalue] = None
                self.c_f[c] = cellvalue
            c+=1
        self.totalcolumns=c
        self.vmr_features['host']=self.vmr_hostfeatures;self.vmr_features['guest'] = self.vmr_guestfeatures

    @decorator_check_profile_ids
    def get_vmrinfo_fromxsl(self,filerow):
        c=1
        while c<self.totalcolumns:
            if c == 1:
                self.vmr_features[self.c_f[c]]=str(self.ws.cell(row=filerow,column=c).value)
            elif (c>1 and c<=self.hostc):self.vmr_features['host'][self.c_f[c]]=self.ws.cell(row=filerow,column=c).value
            else: self.vmr_features['guest'][self.c_f[c]]=self.ws.cell(row=filerow,column=c).value
            c +=1
        self.vmr_features['host']={k:v for k,v in self.vmr_features['host'].items() if v}
        self.vmr_features['guest'] = {k:v for k,v in self.vmr_features['guest'].items() if v}
        self.vmr_features['guest']['callid']=self.vmr_features['guest']['name']='0'+self.vmr_features[self.c_f[1]]
        self.vmr_features['guest']['uri'] =self.vmr_features[self.c_f[1]]
        if ('passcode' in self.vmr_features['host'].keys()):
            self.vmr_features['host']['uri']=self.vmr_features['guest']['uri']
            self.vmr_features['host']['callid'] =self.vmr_features['guest']['callid']
        else:
            raise ValueError('vmr %s need host pin' %self.vmr_features['Vmr Number'])
        # print(self.vmr_features)
        return (self.vmr_features)



    def add_vmrs(self):
        try:
            htmlcontent = api.get(subject='cospaces', quering={'filter': ('%s' % self.vmr_features['Vmr Number'])})
        except Exception as e:
            logging.error(e)
        if htmlcontent.xpath('//cospaces/@total')[0]=='0': #check if the vmr is exist or not
            try:
                guest_status_code=self.api.post(subject='cospaces',payload=self.vmr_features['guest'])
            except Exception as e:
                raise (e)
            try:
                htmlcontent=api.get(subject='cospaces',quering={'filter': ('%s' %self.vmr_features['Vmr Number'])})
            except Exception as e:
                logging.error(e)
            # print(htmlcontent,htmlcontent.xpath('//cospaces/@total')[0])
            if self.vmr_features['host']['passcode']:
                if htmlcontent.xpath('//cospaces/@total')[0]=='1':
                    hostid=htmlcontent.xpath('//cospaces/cospace/@id')[0]
                    try:
                        status_code=self.api.post(('cospaces/%s/accessmethods' %hostid),self.vmr_features['host'])
                        if status_code!=200:
                            raise ('vmr %s FAILED to create host access method' %self.vmr_features['Vmr Number'] )
                        else:
                            logging.info('vmr %s is done', self.vmr_features['Vmr Number'])
                    except Exception as e:
                        raise (e)
        else:
            raise ValueError('vmr %s already exist' %self.vmr_features['Vmr Number'])




if __name__=='__main__':

    server_ip, login= read_serverinfo()
    api = Get_put_post(server_ip['dbserver'], login['username'], login['password'])
    ws,filerow=open_xslsheet()
    vmr=Add_vmrs(worksheet=ws,api=api)

    for i in range(3,filerow):
        try:
            vmr.get_titles()
            vmr.get_vmrinfo_fromxsl(i)
            vmr.add_vmrs()
        except Exception as error:
            logging.error(error)
            continue





