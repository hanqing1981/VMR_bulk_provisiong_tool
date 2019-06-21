from API import Get_put_post
import openpyxl
from collections import OrderedDict
from profiles_passcode_check import decorator_check_profile_ids


import logging
import configparser

logging.basicConfig(filename="systemlogs.txt",
                    format='levelname:%(levelname)s filename: %(filename)s '
                           'outputNumber: [%(lineno)d]  func: %(funcName)s output msg:  %(message)s'
                           ' - %(asctime)s', datefmt='[%d/%b/%Y %H:%M:%S]', level=logging.INFO)


def read_serverinfo():
    cf = configparser.ConfigParser()
    cf.read("system_init_info.ini")
    server_ip = dict(cf.items("server_ip"))
    login = dict(cf.items("server_login"))
    return (server_ip, login)

def open_xslsheet():
    try:
        wb = openpyxl.load_workbook('./VMR_list.xlsx')
        ws=wb.get_sheet_by_name('del')
    except:
        raise
    filerow=1
    while ws.cell(row=filerow,column=1).value:
        filerow +=1
    return (ws,filerow)


class Del_vmrs():
    def __init__(self,worksheet,api):
        self.ws = worksheet
        self.api=api

    def get_vmrinfo_fromxsl(self,filerow):
        c=1
        self.vmr=str(self.ws.cell(row=filerow,column=c).value)




    def del_vmrs(self):
        try:
            htmlcontent = api.get(subject='cospaces', quering={'filter': ('%s' % self.vmr)})
        except Exception as e:
            print(e)
        if htmlcontent.xpath('//cospaces/@total')[0]=='1': #check if the vmr is exist or not
            self.vmrid=htmlcontent.xpath('//cospaces/cospace/@id')[0]
            try:
                status_code=api.deleltes(subject=('cospaces/%s' %self.vmrid))
                if status_code==200:
                    logging.info('vmr %s has been deleted', self.vmr)

            except:
                raise
        else:
            raise ValueError('vmr %s doesnt exist or has multi ids' %self.vmr)




if __name__=='__main__':
    server_ip, login= read_serverinfo()

    api = Get_put_post(server_ip['dbserver'], login['username'], login['password'])
    ws,filerow=open_xslsheet()
    vmr=Del_vmrs(worksheet=ws,api=api)

    for i in range(2,filerow):
        try:
            vmr.get_vmrinfo_fromxsl(i)
            vmr.del_vmrs()
        except Exception as error:
            logging.error(error)
            continue