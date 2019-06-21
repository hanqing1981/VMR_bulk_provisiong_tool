
from profiles_passcode_check import decorator_check_profile_ids
from API import Get_put_post
import openpyxl
from collections import OrderedDict
import configparser
import logging

logging.basicConfig(filename="systemlogs.txt",
                    format='levelname:%(levelname)s filename: %(filename)s '
                           'outputNumber: [%(lineno)d]  func: %(funcName)s output msg:  %(message)s'
                           ' - %(asctime)s', datefmt='[%d/%b/%Y %H:%M:%S]', level=logging.INFO)


# logging configure - import, logging.basicConfig,add logging

def read_serverinfo():
    cf = configparser.ConfigParser()
    cf.read("system_init_info.ini")
    server_ip = dict(cf.items("server_ip"))
    login = dict(cf.items("server_login"))
    return (server_ip, login)


def open_xslsheet():
    try:
        wb = openpyxl.load_workbook('./VMR_list.xlsx')
        ws=wb.get_sheet_by_name('provision')
    except Exception as error:
        print(error)
    filerow=3
    while ws.cell(row=filerow,column=1).value:
        filerow +=1
    return (ws,filerow)


class Add_vmrs():
    def __init__(self,worksheet,api):
        self.ws = worksheet
        self.api=api

    def get_titles(self):
        c=1
        self.c_f={};self.vmr_features={};self.vmr_hostfeatures={};self.vmr_guestfeatures={};self.c_f[c]={}
        self.c_f=OrderedDict(self.c_f);self.vmr_features=OrderedDict(self.vmr_features)
        self.vmr_hostfeatures=OrderedDict(self.vmr_hostfeatures);self.vmr_guestfeatures=OrderedDict(self.vmr_guestfeatures)
        self.c_f=OrderedDict(self.c_f)
        while self.ws.cell(row=2,column=c).value:
            cellvalue=self.ws.cell(row=2,column=c).value
            self.c_f[c]=cellvalue
            if c==1:
                self.vmr_features['Vmr Number'] = cellvalue
                self.c_f[c] = cellvalue
            elif (cellvalue.startswith('host')):
                self.vmr_hostfeatures[cellvalue[4:]] = None
                self.c_f[c] = cellvalue[4:]
                self.hostc = c
            else:
                self.vmr_guestfeatures[cellvalue] = None
                self.c_f[c] = cellvalue
            c+=1
        self.totalcolumns=c
        self.vmr_features['host']=self.vmr_hostfeatures;self.vmr_features['guest'] = self.vmr_guestfeatures

    @decorator_check_profile_ids
    def get_vmrinfo_fromxsl(self,filerow):
        self.vmr_features['host']={key:None for key in self.vmr_features['host']}
        self.vmr_features['guest'] = {key: None for key in self.vmr_features['guest']}
        c=1
        while c<self.totalcolumns:
            if c == 1:
                self.vmr_features[self.c_f[c]]=str(self.ws.cell(row=filerow,column=c).value)
            elif (c>1 and c<=self.hostc):self.vmr_features['host'][self.c_f[c]]=self.ws.cell(row=filerow,column=c).value
            else: self.vmr_features['guest'][self.c_f[c]]=self.ws.cell(row=filerow,column=c).value
            c +=1
        self.vmr_features['host']={k:v for k,v in self.vmr_features['host'].items() if v}
        self.vmr_features['guest'] = {k:v for k,v in self.vmr_features['guest'].items() if v}
        return (self.vmr_features)



    def modify_vmrs(self):
        try:
            htmlcontent = api.get(subject='cospaces', quering={'filter': ('%s' % self.vmr_features['Vmr Number'])})
        except Exception as e:
            print(e)
        if htmlcontent.xpath('//cospaces/@total')[0]!='1': #check if the vmr is exist or not
            raise ValueError('vmr %s doesnt exist or there are multiple vmrs ' % self.vmr_features['Vmr Number'])
        else:
            guestid=htmlcontent.xpath('//cospaces/cospace/@id')[0]
            guest_result = api.put(subject=('cospaces/%s' % guestid),payload=self.vmr_features['guest'])
            vmr_guest = api.get(subject='cospaces/%s/accessmethods/' %guestid)
            if vmr_guest.xpath('//accessmethods/@total')[0] == '1':
                hostid= vmr_guest.xpath('//accessmethods/accessmethod/@id')[0]
                host_result = self.api.put(('cospaces/%s/accessmethods/%s' %(guestid,hostid)), self.vmr_features['host'])
            elif vmr_guest.xpath('//accessmethods/@total')[0] == '0':
                host_result=200
            else:
                raise ValueError('vmr %s multiple accessmethods exist' %self.vmr_features['Vmr Number'])
        if (host_result==200&guest_result==200):
            logging.info('vmr %s is done',self.vmr_features['Vmr Number'])
        else:
            logging.info('vmr %s: host change result %s, guest change result %s', (self.vmr_features['Vmr Number'],host_result,guest_result))


if __name__=='__main__':

    server_ip, login=read_serverinfo()

    api = Get_put_post(server_ip['dbserver'],login['username'],login['password'])
    ws,filerow=open_xslsheet()
    vmr=Add_vmrs(worksheet=ws,api=api)

    for i in range(3,filerow):
        try:
            vmr.get_titles()
            vmr.get_vmrinfo_fromxsl(i)
            vmr.modify_vmrs()
        except Exception as error:
            logging.error(error)
            continue



