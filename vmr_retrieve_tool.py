from API import Get_put_post
import openpyxl
from collections import OrderedDict
import logging
import configparser

logging.basicConfig(filename="systemlogs.txt",
                    format='levelname:%(levelname)s filename: %(filename)s '
                           'outputNumber: [%(lineno)d]  func: %(funcName)s output msg:  %(message)s'
                           ' - %(asctime)s', datefmt='[%d/%b/%Y %H:%M:%S]', level=logging.INFO)


def read_serverinfo():
    cf = configparser.ConfigParser()
    cf.read("system_init_info.ini")
    server_ip = dict(cf.items("server_ip"))
    login = dict(cf.items("server_login"))
    return (server_ip, login)

def open_xslsheet():
    try:
        wb = openpyxl.load_workbook('./VMR_list.xlsx')
        ws=wb.get_sheet_by_name('retrieve')
    except Exception as error:
        logging.error(error)
    filerow=3
    while ws.cell(row=filerow,column=1).value:
        filerow +=1
    return (wb,ws,filerow)



class Retrieve_vmr():
    def __init__(self,workbook,worksheet,api):
        self.wb=workbook
        self.ws=worksheet
        self.api=api

    def get_titles(self):
        c = 1
        self.c_f = {};self.vmr_features = {};self.vmr_hostfeatures = {};self.vmr_guestfeatures = {};
        self.c_f[c] = {}
        self.c_f = OrderedDict(self.c_f);
        self.vmr_features = OrderedDict(self.vmr_features)
        self.vmr_hostfeatures = OrderedDict(self.vmr_hostfeatures);
        self.vmr_guestfeatures = OrderedDict(self.vmr_guestfeatures)
        self.c_f = OrderedDict(self.c_f)
        while self.ws.cell(row=2, column=c).value:
            cellvalue=self.ws.cell(row=2, column=c).value
            if c == 1:
                self.vmr_features['Vmr Number'] = cellvalue
                self.c_f[c] = cellvalue
            elif (cellvalue.startswith('host')):
                self.vmr_hostfeatures[cellvalue[4:]] = None
                self.c_f[c] = cellvalue[4:]
                self.hostc=c
            else:
                self.vmr_guestfeatures[cellvalue] = None
                self.c_f[c] = cellvalue
            c += 1
        self.totalcolumns = c
        self.vmr_features['host'] = self.vmr_hostfeatures;
        self.vmr_features['guest'] = self.vmr_guestfeatures
        # print(self.vmr_features['guest'])

    def get_vmr_number(self,filerow=None):
        self.vmr = self.ws.cell(row=filerow, column=1).value


    def retrieve_vmr_details(self):
        self.vmr_features['guest']={feature:None for feature in self.vmr_features['guest']}
        self.vmr_features['host'] = {feature: None for feature in self.vmr_features['host']}
        # print(('vmr %s before retrieve' %self.vmr), self.vmr_features)
        self.guest_feature_value={}
        self.host_feature_value={}
        self.vmr_id={}
        self.vmr_html = api.get(subject='cospaces', quering={'filter': ('%s' % self.vmr)})
        if self.vmr_html.xpath('//cospaces/@total')[0] == '0':
            raise ValueError('vmr %s doesnt exist' % self.vmr)
        elif self.vmr_html.xpath('//cospaces/@total')[0] != '1':
            raise ValueError('vmr %s has multiple ids' % self.vmr)
        else:
            self.vmr_id['guestid']= self.vmr_html.xpath('//cospaces/cospace/@id')[0]  # add cospace id into dict

            self.guest_fhtml=api.get(subject='cospaces/%s' % self.vmr_id['guestid'])

            for feature in self.vmr_features['guest'].keys():
                if self.guest_fhtml.xpath('//%s' %feature):
                    self.vmr_features['guest'][feature]=self.guest_fhtml.xpath('//%s' %feature)[0].text

            # print(self.vmr_features['guest'])
            self.vmr_features['guest']['id']=self.vmr_id['guestid']
            vmr_guest = api.get(subject=('cospaces/%s/accessmethods/' % self.vmr_features['guest']['id']))
            if vmr_guest.xpath('//accessmethods/@total')[0] == '1':
                self.vmr_id['hostid'] = vmr_guest.xpath('//accessmethods/accessmethod/@id')[0]
                self.host_fhtml = api.get(subject='cospaces/%s/accessmethods/%s' %(self.vmr_features['guest']['id'],self.vmr_id['hostid']))
                for feature in self.vmr_features['host'].keys():
                    if self.host_fhtml.xpath('//%s' % feature):
                        self.vmr_features['host'][feature]=self.host_fhtml.xpath('//%s' % feature)[0].text
                self.vmr_features['host']['id']=self.vmr_id['hostid']
            elif vmr_guest.xpath('//accessmethods/@total')[0] == '0':
                pass
            else:
                raise ValueError(' %s vmr has more then 1 accessmthods' % self.vmr)

        # print(self.vmr_features)

    def write_values_to_excel(self,filerow=None):
        # print(self.vmr_features)
        self.full_valuesxsl={}
        self.hostfxsl={self.host_fns[k]:v for k,v in self.host_feature_value.items()}
        self.full_valuesxsl=self.hostfxsl
        self.full_valuesxsl.update(self.guest_feature_value)

        if self.vmr_features['host']['id']:
            try:
                for c in range(2,self.hostc+1):
                    self.ws.cell(row=filerow,column=c).value=self.vmr_features['host'][self.c_f[c]]
                for c in range(self.hostc+1,self.totalcolumns):
                    self.ws.cell(row=filerow, column=c).value = self.vmr_features['guest'].get(self.c_f[c],None)
            except Exception as error:
                raise ValueError(error)
            else:
                logging.info('vmr %s is done',self.vmr)
        else:
            try:
                for c in range(self.hostc + 1, self.totalcolumns):
                    self.ws.cell(row=filerow, column=c).value = self.vmr_features['guest'].get(self.c_f[c], None)
            except Exception as error:
                raise ValueError(error)
            else:
                logging.info('vmr %s is done', self.vmr)
        self.wb.save('./VMR_list.xlsx')



if __name__=='__main__':
    server_ip, login= read_serverinfo()

    api = Get_put_post(server_ip['dbserver'], login['username'], login['password'])

    wb,ws,filerow = open_xslsheet()
    vmr =Retrieve_vmr(workbook=wb,worksheet=ws,api=api)
    vmr.get_titles()
    for i in range(3, filerow):
        try:
            vmr.get_vmr_number(i)
            vmr.retrieve_vmr_details()
            vmr.write_values_to_excel(i)
        except Exception as error:
            logging.error(error)





